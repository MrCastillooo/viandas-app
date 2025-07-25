import os
import io
import csv
import uuid
import locale
from flask import Flask, Response, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import date, timedelta
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- CONFIGURACIÓN DE IDIOMA ---
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'Spanish')
    except locale.Error:
        print("Advertencia: No se pudo configurar el idioma a español.")

# --- CONFIGURACIÓN DE LA APP ---
app = Flask(__name__)
SECRET_KEY_FROM_ENV = os.environ.get('SECRET_KEY')
DATABASE_URL_FROM_ENV = os.environ.get('DATABASE_URL')

if not SECRET_KEY_FROM_ENV: raise ValueError("¡ERROR CRÍTICO: La variable de entorno SECRET_KEY no está configurada!")
if not DATABASE_URL_FROM_ENV: raise ValueError("¡ERROR CRÍTICO: La variable de entorno DATABASE_URL no está configurada!")

db_url = DATABASE_URL_FROM_ENV
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config['SECRET_KEY'] = SECRET_KEY_FROM_ENV
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# --- MODELOS (sin cambios) ---
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True); name = db.Column(db.String(100), nullable=False); email = db.Column(db.String(100), unique=True, nullable=False); legajo = db.Column(db.String(20), unique=True, nullable=True); password_hash = db.Column(db.String(128)); sector = db.Column(db.String(50)); role = db.Column(db.String(20), default='empleado')
    @property
    def is_admin(self): return self.role == 'admin'
    def set_password(self, password): self.password_hash = generate_password_hash(password)
    def check_password(self, password): return check_password_hash(self.password_hash, password)

class Menu(db.Model):
    id = db.Column(db.Integer, primary_key=True); menu_date = db.Column(db.Date, nullable=False); menu_type = db.Column(db.String(50), nullable=False); description = db.Column(db.String(200), nullable=False)

class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True); user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False); menu_id = db.Column(db.Integer, db.ForeignKey('menu.id'), nullable=True); order_date = db.Column(db.Date, nullable=False); meal_type = db.Column(db.String(10), nullable=False) 
    user = db.relationship('User', backref=db.backref('orders', lazy=True, cascade="all, delete-orphan")); menu = db.relationship('Menu', backref=db.backref('orders', lazy=True))
    __table_args__ = (db.UniqueConstraint('user_id', 'order_date', name='_user_date_uc'),)

class SystemSetting(db.Model):
    key = db.Column(db.String(50), primary_key=True); value = db.Column(db.String(100), nullable=False)

@app.template_filter('format_es')
def format_date_spanish(dt, format_type='full'):
    dias = {1: "Lunes", 2: "Martes", 3: "Miércoles", 4: "Jueves", 5: "Viernes", 6: "Sábado", 7: "Domingo"}; meses = {1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio", 7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"}
    if format_type == 'full': return f"{dias[dt.isoweekday()]}, {dt.day} de {meses[dt.month]} de {dt.year}"
    elif format_type == 'abbr': return f"{dias[dt.isoweekday()][:3]} {dt.day}"
    return dt.strftime('%Y-%m-%d')

@login_manager.user_loader
def load_user(user_id): return User.query.get(int(user_id))

def get_active_week():
    setting = SystemSetting.query.get('week_start_date'); start_date = date.fromisoformat(setting.value) if setting else date(2025, 6, 23)
    return [start_date + timedelta(days=i) for i in range(7)]

# --- RUTAS REESTRUCTURADAS ---

@app.route('/')
def index():
    # La página de inicio ahora simplemente redirige al login si no estás autenticado
    if current_user.is_authenticated:
        if current_user.role == 'admin': return redirect(url_for('admin_summary'))
        if current_user.role == 'encargado': return redirect(url_for('manager_dashboard'))
        if current_user.role == 'empleado': return redirect(url_for('employee_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        user = User.query.filter_by(email=request.form.get('email')).first()
        if user and user.check_password(request.form.get('password')):
            # Forzamos la sesión con remember=True
            login_user(user, remember=True)
            flash(f'Bienvenido, {user.name}!', 'success')
            # Redirección directa al dashboard correspondiente
            if user.role == 'admin': return redirect(url_for('admin_summary'))
            if user.role == 'encargado': return redirect(url_for('manager_dashboard'))
            if user.role == 'empleado': return redirect(url_for('employee_dashboard'))
        else:
            flash('Email o contraseña incorrectos.')
            return redirect(url_for('login'))
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user(); flash('Has cerrado sesión exitosamente.'); return redirect(url_for('login'))

# ... El resto de las rutas no cambian ...
@app.route('/dashboard/empleado')
@login_required
def employee_dashboard():
    # ...
    # El contenido de esta función y todas las demás es el mismo que antes
    # ...
    if current_user.role != 'empleado': return redirect(url_for('index'))
    week_dates = get_active_week(); start_date, end_date = week_dates[0], week_dates[-1]
    menus_query = Menu.query.filter(Menu.menu_date.between(start_date, end_date)).all()
    weekly_menu = {day: [m for m in menus_query if m.menu_date == day] for day in week_dates}
    orders_query = Order.query.filter(Order.user_id == current_user.id, Order.order_date.between(start_date, end_date)).all()
    existing_orders = {order.order_date: order for order in orders_query}
    return render_template('selection_template.html', weekly_menu=weekly_menu, week_dates=week_dates, existing_orders=existing_orders, employee=current_user, title=f"Tus Pedidos de la Semana, {current_user.name}", save_url=url_for('save_week'))

def process_week_selection(user_id):
    week_dates = get_active_week()
    for day in week_dates:
        day_str = day.isoformat(); selection_type = request.form.get(f'selection_type-{day_str}'); existing_order = Order.query.filter_by(user_id=user_id, order_date=day).first()
        if selection_type == 'franco':
            if existing_order: existing_order.menu_id = None; existing_order.meal_type = 'Franco'
            else: db.session.add(Order(user_id=user_id, menu_id=None, order_date=day, meal_type='Franco'))
        elif selection_type == 'pedido':
            selected_dish_id = request.form.get(f'dish-{day_str}'); selected_meal_type = request.form.get(f'meal_type-{day_str}')
            if selected_dish_id and selected_meal_type:
                if existing_order: existing_order.menu_id = int(selected_dish_id); existing_order.meal_type = selected_meal_type
                else: db.session.add(Order(user_id=user_id, menu_id=int(selected_dish_id), order_date=day, meal_type=selected_meal_type))
            elif existing_order: db.session.delete(existing_order)
        else:
            if existing_order: db.session.delete(existing_order)
    db.session.commit()

@app.route('/save_week', methods=['POST'])
@login_required
def save_week():
    if current_user.role != 'empleado': return redirect(url_for('index'))
    process_week_selection(current_user.id); flash('Tus selecciones de la semana han sido guardadas.'); return redirect(url_for('employee_dashboard'))

@app.route('/manager')
@login_required
def manager_dashboard():
    if current_user.role != 'encargado': return redirect(url_for('index'))
    employees = User.query.filter_by(sector=current_user.sector, role='empleado').order_by(User.name).all()
    return render_template('manager_dashboard.html', employees=employees)

@app.route('/manager/select_meals/<int:employee_id>')
@login_required
def manager_select_meals(employee_id):
    if current_user.role != 'encargado': return redirect(url_for('index'))
    employee = User.query.get_or_404(employee_id)
    if employee.sector != current_user.sector: flash('Acceso no autorizado.'); return redirect(url_for('manager_dashboard'))
    week_dates = get_active_week(); start_date, end_date = week_dates[0], week_dates[-1]
    menus_query = Menu.query.filter(Menu.menu_date.between(start_date, end_date)).all()
    weekly_menu = {day: [m for m in menus_query if m.menu_date == day] for day in week_dates}
    orders_query = Order.query.filter(Order.user_id == employee.id, Order.order_date.between(start_date, end_date)).all()
    existing_orders = {order.order_date: order for order in orders_query}
    return render_template('selection_template.html', weekly_menu=weekly_menu, week_dates=week_dates, existing_orders=existing_orders, employee=employee, title=f"Pedidos para {employee.name}", save_url=url_for('save_employee_week', employee_id=employee.id))

@app.route('/manager/save_employee_week/<int:employee_id>', methods=['POST'])
@login_required
def save_employee_week(employee_id):
    if current_user.role != 'encargado': return redirect(url_for('index'))
    employee = User.query.get_or_404(employee_id)
    if employee.sector != current_user.sector: flash('Acceso no autorizado.'); return redirect(url_for('manager_dashboard'))
    process_week_selection(employee_id); flash(f'Selecciones para {employee.name} guardadas.'); return redirect(url_for('manager_select_meals', employee_id=employee.id))

@app.route('/manager/personnel')
@login_required
def manager_personnel():
    if current_user.role != 'encargado': return redirect(url_for('index'))
    employees = User.query.filter_by(sector=current_user.sector, role='empleado').order_by(User.name).all()
    return render_template('manager_personnel.html', employees=employees)

@app.route('/manager/add_employee', methods=['GET', 'POST'])
@login_required
def add_employee():
    if current_user.role != 'encargado': return redirect(url_for('index'))
    if request.method == 'POST':
        name = request.form.get('name'); email = request.form.get('email'); legajo = request.form.get('legajo'); password = request.form.get('password')
        if User.query.filter_by(email=email).first(): flash('El email ya está en uso.'); return redirect(url_for('add_employee'))
        if legajo and User.query.filter_by(legajo=legajo).first(): flash('El número de legajo ya está en uso.'); return redirect(url_for('add_employee'))
        new_user = User(name=name, email=email, legajo=legajo, sector=current_user.sector, role='empleado'); new_user.set_password(password)
        db.session.add(new_user); db.session.commit(); flash(f'Empleado {name} agregado con éxito.'); return redirect(url_for('manager_personnel'))
    return render_template('manager_employee_form.html', title="Agregar Empleado", employee=None)

@app.route('/manager/edit_employee/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    if current_user.role != 'encargado': return redirect(url_for('index'))
    employee = User.query.get_or_404(employee_id)
    if employee.sector != current_user.sector: flash('Acceso no autorizado.'); return redirect(url_for('manager_personnel'))
    if request.method == 'POST':
        new_email = request.form.get('email'); new_legajo = request.form.get('legajo')
        if new_email != employee.email and User.query.filter_by(email=new_email).first():
            flash('El nuevo email ya está en uso por otro usuario.'); return render_template('manager_employee_form.html', title="Editar Empleado", employee=employee)
        if new_legajo and new_legajo != employee.legajo and User.query.filter_by(legajo=new_legajo).first():
            flash('El nuevo legajo ya está en uso por otro usuario.'); return render_template('manager_employee_form.html', title="Editar Empleado", employee=employee)
        employee.name = request.form.get('name'); employee.email = new_email; employee.legajo = new_legajo; password = request.form.get('password')
        if password: employee.set_password(password)
        db.session.commit(); flash(f'Empleado {employee.name} actualizado.'); return redirect(url_for('manager_personnel'))
    return render_template('manager_employee_form.html', title="Editar Empleado", employee=employee)

@app.route('/manager/delete_employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee(employee_id):
    if current_user.role != 'encargado': return redirect(url_for('index'))
    employee = User.query.get_or_404(employee_id)
    if employee.sector != current_user.sector: flash('Acceso no autorizado.'); return redirect(url_for('manager_personnel'))
    flash(f'Empleado {employee.name} eliminado.'); db.session.delete(employee); db.session.commit(); return redirect(url_for('manager_personnel'))

@app.route('/admin/report')
@app.route('/admin/report/<report_date_str>')
@login_required
def admin_dashboard(report_date_str=None):
    if not current_user.is_admin: return redirect(url_for('index')); week_dates = get_active_week(); report_date = date.fromisoformat(report_date_str) if report_date_str else week_dates[0]
    orders_query = db.session.query(Order, User, Menu).join(User, Order.user_id == User.id).outerjoin(Menu, Order.menu_id == Menu.id).filter(Order.order_date == report_date).order_by(User.sector, User.name).all()
    return render_template('admin.html', orders=orders_query, report_date=report_date, week_dates=week_dates)

@app.route('/admin/summary')
@app.route('/admin/summary/<report_date_str>')
@login_required
def admin_summary(report_date_str=None):
    if not current_user.is_admin: return redirect(url_for('index')); week_dates = get_active_week(); report_date = date.fromisoformat(report_date_str) if report_date_str else week_dates[0]
    def get_summary(meal_type): return db.session.query(Menu.menu_type, Menu.description, func.count(Order.id).label('total')).join(Order).filter(Order.order_date == report_date, Order.meal_type == meal_type).group_by(Menu.menu_type, Menu.description).order_by(func.count(Order.id).desc()).all()
    lunch_summary = get_summary('Almuerzo'); dinner_summary = get_summary('Cena'); total_lunch = sum(item.total for item in lunch_summary); total_dinner = sum(item.total for item in dinner_summary)
    return render_template('admin_summary.html', lunch_summary=lunch_summary, dinner_summary=dinner_summary, total_lunch=total_lunch, total_dinner=total_dinner, report_date=report_date, week_dates=week_dates)

@app.route('/admin/edit_order/<int:order_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_order(order_id):
    if not current_user.is_admin: return redirect(url_for('index')); order = Order.query.get_or_404(order_id); available_dishes = Menu.query.filter_by(menu_date=order.order_date).all()
    if request.method == 'POST':
        selection_type = request.form.get('selection_type')
        if selection_type == 'franco': order.menu_id = None; order.meal_type = 'Franco'
        elif selection_type == 'pedido':
            menu_id = request.form.get('dish'); meal_type = request.form.get('meal_type')
            if not menu_id or not meal_type: flash('Si selecciona "Pedido", debe elegir un plato y un tipo de comida.', 'danger'); return redirect(url_for('admin_edit_order', order_id=order_id))
            order.menu_id = int(menu_id); order.meal_type = meal_type
        db.session.commit(); flash(f'El pedido para {order.user.name} ha sido actualizado.', 'success')
        return redirect(url_for('admin_dashboard', report_date_str=order.order_date.isoformat()))
    return render_template('admin_edit_order.html', order=order, available_dishes=available_dishes)

@app.route('/export_excel/<report_date_str>')
@login_required
def export_excel(report_date_str):
    if not current_user.is_admin: return redirect(url_for('index')); report_date = date.fromisoformat(report_date_str)
    query_results = db.session.query(User, Order, Menu).select_from(User).outerjoin(Order, (User.id == Order.user_id) & (Order.order_date == report_date)).outerjoin(Menu, Order.menu_id == Menu.id).filter(User.role == 'empleado').order_by(User.sector, User.name).all()
    wb = Workbook(); ws = wb.active; ws.title = f"Reporte {report_date_str}"
    font_bold = Font(bold=True); font_sector = Font(bold=True, size=14); font_header = Font(name='Calibri', size=11, bold=True); font_data = Font(name='Calibri', size=10); alignment_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True); thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    type_mapping = {"Clásico": "C", "Ensalada": "E", "Tradicional": "T", "Regional": "R", "Sin Gluten": "SG", "Vegetariano": "V", "Ejecutivo Saludable": "ES", "Postre": "P", "Dieta": "D"}
    current_row = 1; current_sector = None
    for user, order, menu in query_results:
        if user.sector != current_sector:
            if current_row > 1: current_row += 2
            ws.cell(row=current_row, column=1, value=f"SECTOR: {user.sector.upper()}").font = font_sector; ws.cell(row=current_row, column=8, value=f"FECHA: {report_date.strftime('%d/%m/%Y')}").font = font_bold; current_row += 2
            headers = ['Empleado', 'Legajo', 'Tipo', 'Estado', 'Turno', 'Hora de ingreso', 'Hora de egreso', 'Firma del Empleado', 'Firma del referente']
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=i, value=header); cell.font = font_header; cell.border = thin_border; cell.alignment = alignment_center_wrap
            current_sector = user.sector; current_row += 1
        tipo_code = ''; estado = 'Sin Pedido'
        if order:
            estado = order.meal_type
            if order.meal_type != 'Franco' and menu: tipo_code = type_mapping.get(menu.menu_type, '?')
        data_row = [user.name, user.legajo or '', tipo_code, estado, '', '', '', '', '']
        for i, value in enumerate(data_row, 1):
            cell = ws.cell(row=current_row, column=i, value=value); cell.border = thin_border; cell.font = font_data; cell.alignment = alignment_center_wrap
        ws.row_dimensions[current_row].height = 40; current_row += 1
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 10; ws.column_dimensions['C'].width = 8; ws.column_dimensions['D'].width = 12; ws.column_dimensions['E'].width = 12; ws.column_dimensions['F'].width = 15; ws.column_dimensions['G'].width = 15; ws.column_dimensions['H'].width = 25; ws.column_dimensions['I'].width = 25
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE; ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0 
    in_memory_fp = io.BytesIO(); wb.save(in_memory_fp); in_memory_fp.seek(0) 
    return Response(in_memory_fp, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-disposition": f"attachment; filename=Reporte_Viandas_{report_date_str}.xlsx"})

@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
def admin_settings():
    if not current_user.is_admin: return redirect(url_for('index'));
    if request.method == 'POST':
        setting = SystemSetting.query.get('week_start_date'); start_date_str = request.form.get('week_start_date')
        if setting: setting.value = start_date_str
        else: db.session.add(SystemSetting(key='week_start_date', value=start_date_str))
        db.session.commit(); flash('La fecha de inicio de la semana ha sido actualizada.'); return redirect(url_for('admin_settings'))
    current_setting = SystemSetting.query.get('week_start_date'); return render_template('admin_settings.html', current_setting=current_setting)

@app.route('/admin/manage_menu')
@login_required
def admin_manage_menu():
    if not current_user.is_admin: return redirect(url_for('index'))
    week_dates = get_active_week(); start_date, end_date = week_dates[0], week_dates[-1]; menus_query = Menu.query.filter(Menu.menu_date.between(start_date, end_date)).order_by(Menu.menu_type).all()
    grouped_menus = {day: [] for day in week_dates}
    for menu in menus_query:
        if menu.menu_date in grouped_menus: grouped_menus[menu.menu_date].append(menu)
    return render_template('admin_manage_menu.html', grouped_menus=grouped_menus, week_dates=week_dates)

@app.route('/admin/edit_menu_item/<int:menu_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_menu_item(menu_id):
    if not current_user.is_admin: return redirect(url_for('index'))
    menu_item = Menu.query.get_or_404(menu_id)
    if request.method == 'POST':
        menu_item.description = request.form.get('description'); menu_item.menu_type = request.form.get('menu_type'); db.session.commit()
        flash('El plato ha sido actualizado con éxito.'); return redirect(url_for('admin_manage_menu'))
    return render_template('admin_edit_menu_item.html', menu_item=menu_item)

@app.cli.command("init-db")
def init_db_command():
    db.drop_all(); db.create_all()
    EMAIL_DOMAIN = "tudominio.com"
    db.session.add(SystemSetting(key='week_start_date', value='2025-06-23'))
    admin_user = User(name='Super Admin', email=f"admin@{EMAIL_DOMAIN}", role='admin', sector='Gerencia', legajo='001'); admin_user.set_password('admin123')
    db.session.add(admin_user)
    sectores = [ "Administración", "MKT", "ATC", "Cajas", "Gastronomia", "Limpieza", "Mantenimiento", "Monitoreo", "RRHH", "Sala", "Seguridad", "Sistemas", "Slot", "Tesoreria", "Cardenales S.A.S" ]
    for sector in sectores:
        email_sector = sector.lower(); replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', ' ': '', '.': ''};
        for old, new in replacements.items(): email_sector = email_sector.replace(old, new)
        encargado = User(name=f"Encargado {sector}", email=f"encargado.{email_sector}@{EMAIL_DOMAIN}", role='encargado', sector=sector); encargado.set_password('encargado123')
        db.session.add(encargado)
    empleado1 = User(name='Jose Valencia', email=f"jose.v@{EMAIL_DOMAIN}", role='empleado', sector='Administración', legajo='1001'); empleado1.set_password('jose123')
    empleado2 = User(name='Fernando Arapa', email=f"fernando.a@{EMAIL_DOMAIN}", role='empleado', sector='Administración', legajo='1002'); empleado2.set_password('fer123')
    empleado3 = User(name='Lucas Siarez', email=f"lucas.s@{EMAIL_DOMAIN}", role='empleado', sector='Seguridad', legajo='2001'); empleado3.set_password('lucas123')
    db.session.add_all([empleado1, empleado2, empleado3])
    menu_data = { date(2025, 6, 23):[{"type":"Clásico","desc":"MILANESA DE TERNERA CON PURE"},{"type":"Ensalada","desc":"ENSALADA COMPLETA"},{"type":"Tradicional","desc":'PUCHERO "SALPICON ANDINO"'},{"type":"Regional","desc":"POLLO A LA CREMA CON PAPAS"},{"type":"Sin Gluten","desc":"MEDALLONES DE GARBANZO Y CALABAZA CON ENS TIBIA"},{"type":"Vegetariano","desc":"LASAGNA DE CHOCLO Y QUESO"},{"type":"Ejecutivo Saludable","desc":"ZAPALLITOS RELLENOS CON PURE MIXTO"},{"type":"Postre","desc":"FRUTA FRESCA"},{"type":"Dieta","desc":"PATA MUSLO HERVIDA CON REVUELTOS DE ZAPALLITOS"}], date(2025, 6, 24):[{"type":"Clásico","desc":"MILANESA DE TERNERA CON PURE"},{"type":"Ensalada","desc":"ENSALADA DE LEGUMBRES"},{"type":"Tradicional","desc":"TALLERINES AL HUEVO CON TUCO DE POLLO Y QUESO SARDO"},{"type":"Regional","desc":"PIZZA BALCARSE"},{"type":"Sin Gluten","desc":"COCIDO DE LENTEJAS"},{"type":"Vegetariano","desc":"BOMBAS DE PAPA Y QUESO CON ENS TIBIA"},{"type":"Ejecutivo Saludable","desc":"POLLO CON VEGETALES ESTOFADOS"},{"type":"Postre","desc":"GELATINA"},{"type":"Dieta","desc":"BIFE DE TERNERA CON VERDURA COCIDAS"}], date(2025, 6, 25):[{"type":"Clásico","desc":"MILANESA DE TERNERA CON PURE"},{"type":"Ensalada","desc":"ENSALADA FRESCA"},{"type":"Tradicional","desc":"BOMBAS DE PAPA Y CARNE CON ENS TIBIA"},{"type":"Regional","desc":"POLLO AL VINO BLANCO CON PURE"},{"type":"Sin Gluten","desc":"RAVIOLES VEGANOS EN SALSA DE HONGOS"},{"type":"Vegetariano","desc":"RISOTO DE QUINOA Y CHAMPIGNONES CON ENS TIBIA"},{"type":"Ejecutivo Saludable","desc":"GUISO DE VERDURAS Y FIDEO MONITOS"},{"type":"Postre","desc":"TORTA DE CHOCOLATE"},{"type":"Dieta","desc":"BIFE DE POLLO CON FIDEOS CABELLO DE ANGEL"}], date(2025, 6, 26):[{"type":"Clásico","desc":"MILANESA DE TERNERA CON PURE"},{"type":"Ensalada","desc":"ENSALADA COCIDA"},{"type":"Tradicional","desc":"SUPREMA NAPOLITANA CON ARROZ SABORIZADO"},{"type":"Regional","desc":"COSTELETA DE CARNE CON PAPAS CRIOLLITA"},{"type":"Sin Gluten","desc":"COCIDO DE GARBANZOS"},{"type":"Vegetariano","desc":"CALZON DE VERDURAS, QUESO Y HUEVO CON ENS"},{"type":"Ejecutivo Saludable","desc":"MERLUZA GRATINADA CON PURE MIXTO"},{"type":"Postre","desc":"FRUTA FRESCA"},{"type":"Dieta","desc":"PATA MUSLO HERVIDA CON ARROZ BLANCO + QUESO"}], date(2025, 6, 27):[{"type":"Clásico","desc":"MILANESA DE TERNERA CON PURE"},{"type":"Ensalada","desc":"ENSALADA WARA"},{"type":"Tradicional","desc":"COSTILLITAS ESTOFADAS CON ARROZ AMARILLO"},{"type":"Regional","desc":"MARINERA DE TERNERA CON FIDEOS TRICOLOR"},{"type":"Sin Gluten","desc":"ÑOQUIS DE PAPA Y BATATA CON SALSA TEXTURIZADA"},{"type":"Vegetariano","desc":"EMPANADAS DE CHOCLO Y QUESO"},{"type":"Ejecutivo Saludable","desc":"CAZUELA DE POLLO Y VEGETALES"},{"type":"Postre","desc":"POSTRE MARACUYA"},{"type":"Dieta","desc":"MERLUZA AL VAPOR CON PURE DE CALABAZA"}], date(2025, 6, 28):[{"type":"Clásico","desc":"MILANESA DE TERNERA CON PURE"},{"type":"Ensalada","desc":"ENSALADA KETO"},{"type":"Tradicional","desc":"HAMBURGUESA PATY CON PAPAS"},{"type":"Vegetariano","desc":"CANELONES DE RICOTA Y ESPINACA CON SALSA MIXTA Y QUESO SARDO"},{"type":"Ejecutivo Saludable","desc":"ROLLITO DE JAMON Y QUESO CON ENS TIBIA"},{"type":"Postre","desc":"FLAN CON DULCE DE LECHE"},{"type":"Dieta","desc":"BIFE DE TERNERA CON REVUELTO DE ZAPALLITOS"}], date(2025, 6, 29):[{"type":"Clásico","desc":"MILANESA DE TERNERA CON PURE"},{"type":"Ensalada","desc":"ENSALADA ANDINA"},{"type":"Tradicional","desc":"ÑOQUIS DE PAPA CON TUCO DE TERNERA"},{"type":"Vegetariano","desc":"PIZZA MIXTA (FUGAZETA Y HUEVO)"},{"type":"Ejecutivo Saludable","desc":"PATA MUSLO RELLENA CON COCHON DE VEGETALES"},{"type":"Postre","desc":"DURAZNO NATURAL"},{"type":"Dieta","desc":"PECHUGA HERVIDA CON VEGETALES COCIDOS"}]}
    for day, menus in menu_data.items():
        for menu_item in menus: db.session.add(Menu(menu_date=day, menu_type=menu_item["type"], description=menu_item["desc"]))
    db.session.commit()
    print(f"Base de datos inicializada con el dominio de email: {EMAIL_DOMAIN}")